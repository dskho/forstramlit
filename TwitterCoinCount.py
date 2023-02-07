import tweepy
import openpyxl
import time
import requests
import datetime
#可以正常运行了 持续运行
#添加发送功能
def get_tweets():
    # 设置Twitter API的认证信息


    # 建立Twitter API的认证
    auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
    auth.set_access_token(access_token, access_token_secret)
    api = tweepy.API(auth)

    # 获取所关注的人的最新推文
    public_tweets = api.home_timeline()

    # 保存推文内容
    tweets = []
    for tweet in public_tweets:
        print(f"Tweet: {tweet.text}")
        tweets.append(tweet.text)

    return tweets


def count_words(tweets):
    # 建立一个字典，用于保存每个单词的出现次数
    word_count = {}

    # 遍历每个推文
    for tweet in tweets:
        # 分词
        words = tweet.split()
        # 遍历每个单词
        for word in words:
            # 统计以 $ 开头的单词的数量
            if word.startswith('$') or word.startswith("#") or word.startswith("@"):
                if word in word_count:
                    word_count[word] += 1
                else:
                    word_count[word] = 1

    return word_count

'''
def save_to_excel(word_count):
    # 建立一个excel文件并保存统计的结果
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Word Count"
    sheet['A1'] = "Word"
    sheet['B1'] = "Count"
    row = 2
    for word, count in word_count.items():
        sheet.cell(row=row, column=1, value=word)
        sheet.cell(row=row, column=2, value=count)
        row += 1
    wb.save("results.xlsx")
    print("Results saved to results.xlsx")
'''

def save_to_excel(word_count):
    # 建立一个excel文件并保存统计的结果
    wb = openpyxl.load_workbook("results.xlsx")
    sheet = wb.active
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    for word, count in word_count.items():
        sheet.cell(row=row, column=2, value=word)
        sheet.cell(row=row, column=3, value=count)
    wb.save("results.xlsx")
    print("Results saved to results.xlsx")

def notify_telegram(word_count):
    # replace <BOT_TOKEN> and <CHAT_ID> with your own values


    # 发送信息到telegram的代码
    # 设置请求头和请求参数
    headers = {"Content-Type": "application/json"}
    params = {"chat_id": CHAT_ID, "text": "过去半小时内关注的twitter上提到的加密货币次数: " + str(word_count)}
    # 向telegram的bot发送消息
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    requests.post(url, headers=headers, json=params)

if __name__ == '__main__':
    while True:
        tweets = get_tweets()
        word_count = count_words(tweets)
        for word, count in word_count.items():
            print("{}: {}".format(word, count))

        save_to_excel(word_count)
        notify_telegram(word_count)
        time.sleep(1800)
